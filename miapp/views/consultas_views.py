from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db.models import (
    Q, F, Sum, Count, Case, When, Value, CharField, DecimalField
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from ..models import (
    MovimientoInventario,
    TipoMovimiento,
    Bodega,
    Proveedor,
    Producto,
    Compra,
    Pedido,
    EstadoPedido,
    EstadoLote,
    TransferenciaInterna,
    User,DetallePedido
)
from .helpers import _get_context_base
@login_required
def kardex_consulta(request):
    """
    Kardex Dinámico Omnicanal: 
    Rastrea el saldo acumulado real considerando despachos fraccionados.
    """
    ctx = _get_context_base(request)
    
    # 1. CAPTURA DE FILTROS
    q = request.GET.get("q", "").strip()
    bodega_id = request.GET.get("bodega", "")
    desde = request.GET.get("desde", "")
    hasta = request.GET.get("hasta", "")
    
    # 2. CONSULTA BASE OPTIMIZADA
    # Incluimos 'pedido' y 'bodega_origen/destino' para ver el flujo multi-bodega
    movimientos_qs = MovimientoInventario.objects.select_related(
        'tipo_movimiento', 'usuario', 'stock_lote__producto', 
        'inventario__bodega', 'pedido', 'bodega_origen', 'bodega_destino'
    ).order_by('fecha', 'id') 

    if q:
        movimientos_qs = movimientos_qs.filter(
            Q(stock_lote__producto__nombre__icontains=q) | 
            Q(stock_lote__lote__icontains=q) |
            Q(pedido__codigo__icontains=q)
        )
    if bodega_id:
        movimientos_qs = movimientos_qs.filter(inventario__bodega_id=bodega_id)
    if desde:
        movimientos_qs = movimientos_qs.filter(fecha__date__gte=desde)
    if hasta:
        movimientos_qs = movimientos_qs.filter(fecha__date__lte=hasta)

    # 3. LÓGICA DE SALDO ACUMULADO
    saldos_temporales = {}
    movimientos_con_saldo = []

    for mov in movimientos_qs:
        # La clave sigue siendo Producto + Bodega para el saldo físico local
        key = (mov.stock_lote.producto_id, mov.inventario.bodega_id)
        saldo_anterior = saldos_temporales.get(key, 0)
        
        variacion = mov.cantidad if mov.tipo_movimiento.es_entrada else -mov.cantidad
        saldo_actual = saldo_anterior + variacion
        
        # Inyectamos datos extra para la auditoría de fraccionamiento
        mov.saldo_calculado = saldo_actual
        mov.saldo_anterior = saldo_anterior
        
        saldos_temporales[key] = saldo_actual
        movimientos_con_saldo.append(mov)

    movimientos_con_saldo.reverse()

    # KPIs de la consulta (Valorización de lo filtrado)
    resumen_kpi = {
        "total_valor_salida": sum(m.valor_total for m in movimientos_con_saldo if not m.tipo_movimiento.es_entrada),
        "total_unidades_movidas": sum(m.cantidad for m in movimientos_con_saldo)
    }

    ctx.update({
        "movimientos": movimientos_con_saldo,
        "resumen_kpi": resumen_kpi,
        "bodegas": Bodega.objects.filter(es_activo=True),
        "filtros": request.GET,
        "titulo": "Auditoría de Kardex Multi-Bodega"
    })
    
    return render(request, "miapp/consultas/kardex_list.html", ctx)

@login_required
def kardex_detalle(request, pk):
    """
    Vista detallada de auditoría forense.
    """
    ctx = _get_context_base(request)
    
    movimiento = get_object_or_404(
        MovimientoInventario.objects.select_related(
            'tipo_movimiento', 'usuario', 'stock_lote__producto__perfil_caducidad', 
            'stock_lote__proveedor', 'bodega_origen', 'bodega_destino',
            'pedido__cliente', 'pedido__usuario',
            'stock_lote__ubicacion__rack__zona__bodega'
        ), pk=pk
    )

    arbol_genealogico = movimiento.stock_lote.obtener_arbol_genealogico()

    rentabilidad = None
    if movimiento.pedido:
        detalle_p = DetallePedido.objects.filter(
            pedido=movimiento.pedido, 
            producto=movimiento.stock_lote.producto
        ).first()
        
        if detalle_p:
            precio_venta = detalle_p.precio_unitario
            costo_lote = movimiento.valor_unitario 
            rentabilidad = {
                'precio_vta': precio_venta,
                'margen_unitario': precio_venta - costo_lote,
                'ganancia_total': (precio_venta - costo_lote) * movimiento.cantidad,
                'porcentaje': ((precio_venta - costo_lote) / precio_venta * 100) if precio_venta > 0 else 0
            }

    ctx.update({
        "mov": movimiento,
        "arbol": arbol_genealogico,
        "rentabilidad": rentabilidad,
        "titulo": f"Auditoría Forense: {movimiento.codigo}"
    })
    return render(request, "miapp/consultas/kardex_detalle.html", ctx)

# --- REEMPLAZA TU FUNCIÓN ACTUAL POR ESTA ---
@login_required
def transferencias_consulta(request):
    ctx = _get_context_base(request)
    
    # 1. Filtros
    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "")
    bodega_id = request.GET.get("bodega", "")
    desde = request.GET.get("desde", "")
    hasta = request.GET.get("hasta", "")
    tipo_trf = request.GET.get("tipo", "")

    # 2. CONSULTA OPTIMIZADA (Calculamos el dinero directo en la Base de Datos)
    items = TransferenciaInterna.objects.select_related(
        'bodega_origen', 'bodega_destino', 'usuario_solicita', 'usuario_ejecuta'
    ).annotate(
        # Aquí sumamos: Cantidad * Costo del lote original
        valor_auditado=Coalesce(
            Sum(
                F('detalles__lotes_asignados__cantidad') * F('detalles__lotes_asignados__stock_lote_origen__costo_compra_lote'),
                output_field=DecimalField()
            ),
            Value(0, output_field=DecimalField())
        )
    ).order_by('-fecha_creacion')

    # 3. Aplicar Filtros
    if q: items = items.filter(Q(codigo__icontains=q) | Q(observaciones__icontains=q))
    if estado: items = items.filter(estado=estado)
    if tipo_trf: items = items.filter(tipo=tipo_trf)
    if bodega_id: items = items.filter(Q(bodega_origen_id=bodega_id) | Q(bodega_destino_id=bodega_id))
    if desde: items = items.filter(fecha_creacion__date__gte=desde)
    if hasta: items = items.filter(fecha_creacion__date__lte=hasta)

    # 4. KPIs Rápidos
    lista_items = list(items) # Convertimos a lista una sola vez
    
    # Sumamos la columna que calculó la base de datos
    valor_total_consulta = sum((x.valor_auditado or 0) for x in lista_items)

    ctx.update({
        "items": lista_items, # Pasamos la lista ya procesada
        "kpis": {
            "pendientes": sum(1 for x in lista_items if x.estado in ['SOLICITADA', 'EN_PREPARACION']),
            "en_transito": sum(1 for x in lista_items if x.estado == 'DESPACHADA'),
            "valor_total": valor_total_consulta
        },
        "bodegas": Bodega.objects.filter(es_activo=True),
        "estados": TransferenciaInterna.Estado.choices,
        "tipos": TransferenciaInterna.Tipo.choices,
        "filtros": request.GET,
        "titulo": "Monitor de Transferencias"
    })
    return render(request, "miapp/consultas/transferencia_list.html", ctx)


@login_required
def transferencia_detalle(request, pk):
    ctx = _get_context_base(request)
    # Traemos la transferencia con prefetch de sus lotes para ver qué viene en el camión
    trf = get_object_or_404(
        TransferenciaInterna.objects.select_related(
            'bodega_origen', 'bodega_destino', 'usuario_solicita', 'usuario_ejecuta', 'pedido'
        ).prefetch_related(
            'detalles__producto',
            'detalles__lotes_asignados__stock_lote_origen',
            'detalles__lotes_asignados__ubicacion_destino'
        ), pk=pk
    )

    # Calculamos el valor monetario de este camión específico
    valor_camion = trf.obtener_valor_en_transito()

    ctx.update({
        "trf": trf,
        "valor_camion": valor_camion,
        "titulo": f"Auditoría de Traspaso: {trf.codigo}"
    })
    return render(request, "miapp/consultas/transferencia_detalle.html", ctx)
@login_required
def pedidos_consulta(request):
    """
    Monitor de Ventas y Gestión Logística:
    Visualiza el flujo de ingresos, rentabilidad y el progreso de despachos fraccionados.
    """
    ctx = _get_context_base(request)
    
    # 1. Captura de Filtros
    q = request.GET.get("q", "").strip()
    estado_id = request.GET.get("estado", "")
    desde = request.GET.get("desde", "")
    hasta = request.GET.get("hasta", "")

    # 2. Consulta Base OPTIMIZADA
    # Se incluye select_related para evitar el problema de consultas N+1 y prefetch_related
    # para acceder eficientemente a los detalles y productos en el resumen de la tabla.
    pedidos = Pedido.objects.select_related(
        'cliente', 'estado_pedido', 'usuario'
    ).prefetch_related(
        'detalles__producto' 
    ).all().order_by('-fecha_creacion')

    # 3. Aplicación de Filtros Dinámicos
    if q:
        pedidos = pedidos.filter(
            Q(codigo__icontains=q) | 
            Q(cliente__nombres__icontains=q) | 
            Q(cliente__apellidos__icontains=q) |
            Q(cliente__numero_identificacion__icontains=q)
        )
    if estado_id:
        pedidos = pedidos.filter(estado_pedido_id=estado_id)
    if desde:
        pedidos = pedidos.filter(fecha_creacion__date__gte=desde)
    if hasta:
        pedidos = pedidos.filter(fecha_creacion__date__lte=hasta)

    # 4. INTELIGENCIA DE NEGOCIO (KPIs)
    # Se calculan métricas financieras globales basadas en los filtros aplicados.
    kpis = pedidos.aggregate(
        total_ventas=Coalesce(Sum('total'), Value(Decimal("0.00"))),
        total_costos=Coalesce(Sum('total_costo'), Value(Decimal("0.00"))),
        total_ganancia=Coalesce(Sum('ganancia_total'), Value(Decimal("0.00"))),
        conteo=Count('id')
    )
    
    # Cálculo del porcentaje de margen global del periodo consultado
    margen_global = 0
    if kpis['total_ventas'] > 0:
        margen_global = (kpis['total_ganancia'] / kpis['total_ventas']) * 100

    # 5. Exportación a Excel Profesional
    if 'export' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Reporte_Ventas_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas y Rentabilidad"
        
        # Encabezados del reporte
        headers = [
            'Código', 'Fecha', 'Cliente', 'Origen', 'Estado', 
            'Subtotal', 'IVA', 'Total Venta', 'Costo Total', 
            'Utilidad', '% Margen', '% Avance Físico'
        ]
        ws.append(headers)
        
        # Estilo para encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for p in pedidos:
            margen_p = (p.ganancia_total / p.total * 100) if p.total > 0 else 0
            ws.append([
                p.codigo, 
                p.fecha_creacion.strftime("%d/%m/%Y"), 
                f"{p.cliente.nombres} {p.cliente.apellidos}",
                p.get_origen_display(),
                p.estado_pedido.descripcion,
                float(p.subtotal), 
                float(p.monto_iva), 
                float(p.total),
                float(p.total_costo), 
                float(p.ganancia_total), 
                f"{margen_p:.2f}%",
                f"{p.porcentaje_completado}%"
            ])
        wb.save(response)
        return response

    # 6. Preparación de Contexto para la Plantilla
    ctx.update({
        "pedidos": pedidos,
        "kpis": kpis,
        "margen_global": margen_global,
        "estados": EstadoPedido.objects.all(),
        "filtros": {"q": q, "estado_id": estado_id, "desde": desde, "hasta": hasta},
        "titulo": "Monitor de Gestión de Pedidos y Ventas"
    })
    
    return render(request, "miapp/consultas/pedidos_list.html", ctx)

@login_required
def pedido_detalle(request, pk):
    """
    Vista detallada del pedido con enfoque en Estado Logístico y Financiero.
    """
    ctx = _get_context_base(request)
    pedido = get_object_or_404(
        Pedido.objects.select_related('cliente', 'usuario', 'estado_pedido')
        .prefetch_related('detalles__producto'), 
        pk=pk
    )
    
    # Calculamos el porcentaje para el cuadro de diagnóstico del Admin
    margen_porcentaje = 0
    if pedido.total > 0:
        margen_porcentaje = (pedido.ganancia_total / pedido.total) * 100

    ctx.update({
        "pedido": pedido,
        "margen_porcentaje": margen_porcentaje,
        "titulo": f"Detalle Pedido: {pedido.codigo}"
    })
    return render(request, "miapp/consultas/pedido_detalle.html", ctx)


@login_required
def compras_list(request):
    # 1. Obtener parámetros
    query = request.GET.get('q')
    proveedor_id = request.GET.get('proveedor')
    bodega_id = request.GET.get('bodega')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # --- NUEVOS FILTROS ---
    producto_id = request.GET.get('producto')
    usuario_id = request.GET.get('usuario')
    monto_min = request.GET.get('monto_min')
    monto_max = request.GET.get('monto_max')

    # 2. Query Base
    compras = Compra.objects.select_related('proveedor', 'usuario', 'bodega_destino').all().order_by('-fecha')

    # 3. Aplicar Filtros
    if query:
        compras = compras.filter(numero_factura__icontains=query)
    if proveedor_id:
        compras = compras.filter(proveedor_id=proveedor_id)
    if bodega_id:
        compras = compras.filter(bodega_destino_id=bodega_id)
    if fecha_inicio:
        compras = compras.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        compras = compras.filter(fecha__date__lte=fecha_fin)

    # --- LÓGICA DE NUEVOS FILTROS ---
    if producto_id:
        # Busca compras que tengan AL MENOS un detalle con ese producto
        compras = compras.filter(detalles__producto_id=producto_id).distinct()
        
    if usuario_id:
        compras = compras.filter(usuario_id=usuario_id)
        
    if monto_min:
        compras = compras.filter(total__gte=monto_min)
        
    if monto_max:
        compras = compras.filter(total__lte=monto_max)

    # 4. Contexto
    ctx = _get_context_base(request)
    
    # Flag para saber si hay filtros avanzados activos (para dejar abierto el acordeón)
    filtros_avanzados = any([producto_id, usuario_id, monto_min, monto_max])

    ctx.update({
        'compras': compras,
        # Mantener valores en inputs
        'query': query,
        'filtro_proveedor': int(proveedor_id) if proveedor_id else '',
        'filtro_bodega': int(bodega_id) if bodega_id else '',
        'filtro_producto': int(producto_id) if producto_id else '',
        'filtro_usuario': int(usuario_id) if usuario_id else '',
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'monto_min': monto_min,
        'monto_max': monto_max,
        'filtros_avanzados': filtros_avanzados,
        
        # Listas para selects
        'proveedores': Proveedor.objects.filter(es_activo=True),
        'bodegas': Bodega.objects.filter(es_activo=True),
        'productos': Producto.objects.filter(es_activo=True), # Para el filtro por producto
        'usuarios': User.objects.filter(is_active=True),      # Para el filtro por usuario
    })
    
    return render(request, "miapp/consultas/compras_list.html", ctx)


@login_required
def compra_detail(request, pk):
    """
    Vista de detalle de una compra específica (Cabecera + Líneas).
    """
    # Obtenemos la compra y optimizamos la consulta de detalles
    compra = get_object_or_404(
        Compra.objects.select_related('proveedor', 'bodega_destino', 'usuario'), 
        pk=pk
    )
    
    # Traemos los detalles con sus productos y lotes generados
    detalles = compra.detalles.select_related('producto', 'lote_generado').all()

    ctx = _get_context_base(request)
    ctx.update({
        'compra': compra,
        'detalles': detalles,
        'titulo': f'Detalle Compra {compra.numero_factura}'
    })
    return render(request, "miapp/consultas/compra_detail.html", ctx)   